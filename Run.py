import os
import sys

from lxml import html
from openpyxl import load_workbook, Workbook

from Response import send_request

def write_from_dict_to_xlsx(dict):
    try:
        filename = 'output.xlsx'
        wb = load_workbook(filename)
        ws = wb.active

        ws.append([
            dict['company_name'],
            dict['fio'],
            dict['phone'],
            dict['email'],
            dict['company_site_url'],
            dict['url'],
        ])

        wb.save(filename)
    except Exception as e:
        print(e)


if __name__ == '__main__':
    output = 'output.xlsx'
    domain = 'hh.ru'
    urls = ['https://ivanovo.hh.ru/search/vacancy?clusters=true&area=1&specialization=29&no_magic=true&enable_snippets=true&salary=&st=searchVacancy&text=%D0%9E%D0%BF%D0%B5%D1%80%D0%B0%D1%82%D0%BE%D1%80+%D1%87%D0%BF%D1%83+%D1%81%D1%82%D0%B0%D0%BD%D0%BA%D0%B0&from=suggest_post']

    if os.path.exists(output) == True:
        os.remove(output)

    wb = Workbook()
    wb.save(output)

    wb = load_workbook(output)
    ws = wb.active
    ws.append([
        'Название компании',
        'Контактное лицо',
        'Телефон',
        'Email',
        'Сайт',
        'Url',
    ])
    wb.save(output)

    for url in urls:
        response = send_request(url)
        list_tree = html.document_fromstring(response.content)

        elements = list_tree.xpath("//a[@data-qa='vacancy-serp__vacancy-title']/@href")

        for page_url in elements:
            data = {
                'fio': '',
                'phone': '',
                'company_name': '',
                'email': '',
                'company_site_url': '',
                'url': page_url,
            }
            page_response = send_request(page_url)
            page_tree = html.document_fromstring(page_response.content)

            data['company_name'] = ''.join(page_tree.xpath("//a[@data-qa='vacancy-company-name']//text()"))
            company_page_link = page_tree.xpath("//a[@data-qa='vacancy-company-name']/@href")[0]

            contacts_info = page_tree.xpath("//div[@class='vacancy-contacts__body']//text()")
            print(contacts_info)

            data['fio'] = ''.join(page_tree.xpath("//p[@data-qa='vacancy-contacts__fio']//text()"))
            data['phone'] = ','.join(page_tree.xpath("//p[@class='vacancy-contacts__phone-desktop']/text()"))
            data['email'] = ','.join(page_tree.xpath("//a[@data-qa='vacancy-contacts__email']/text()"))

            temporary_list = url.split('.' + domain)

            if len(temporary_list) > 1:
                company_page_url = str(temporary_list[0]) + '.' + domain + company_page_link
            else:
                company_page_url = 'https://' + domain + company_page_link

            company_page_response = send_request(company_page_url)
            company_page_tree = html.document_fromstring(company_page_response.content)
            data['company_site_url'] = ','.join(company_page_tree.xpath("//a[@data-qa='sidebar-company-site']/@href"))

            write_from_dict_to_xlsx(data)
            print(data)
            print('==============')
